import requests
from requests_kerberos import HTTPKerberosAuth
from requests.auth import HTTPBasicAuth
import urllib3
import sqlite3
from sqlite3 import Error
import os
import sys
import ybo_cf_db
import pandas as pd
import csv
import re
from datetime import *
import json
import argparse
import urllib3
import logging
import openpyxl as op
import shutil
import math

if not os.path.exists('Logs'):
    os.makedirs('Logs')

#setting configuration of logging
logging.basicConfig(filename='./Logs/ybo_logs.txt', filemode='w', format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%d-%m-%Y %H:%M:%S', level=logging.INFO)
logging.info("Imports done. Initializing variables...")
 
#    Example commands for execution: 
#    py ybo_local.py -p bios.alderlake -y True -s S -c bios -tp bat -f False -t 10 
#    py ybo_local_copy.py -p bios.raptorlake -s S -c bios -tp fv -f True -tpf C:\NAMAN\Share\YBO\rpl-s 5.csv -t 5
#    py ybo_local.py -p bios.raptorlake -s P -c bios -tp bat -f False -t 5

parser = argparse.ArgumentParser()
parser.add_argument('-p', help="Name of program")
# parser.add_argument('-y', help="SKU Present? (True/False)")
parser.add_argument('-s', help="Name of SKU", required=False)
parser.add_argument('-c', help="Name of charter")
parser.add_argument('-tp', help="Name of test plan")
parser.add_argument('-f', help="Test plan to be manually uploaded? (True/False)")
parser.add_argument('-tpf', help="Test plan file path", required=False)
parser.add_argument('-t', help="Threshold", type = int)
# parser.add_argument('-l', help="Mandatory Test case level")
args = parser.parse_args()

logging.info("Inputs received -> ")
logging.info("Input - Name of program = {0}".format(args.p.lower()))
# logging.info("Input - SKU present? (True/False) = {0}".format(args.y))
# if args.y in [True, "True", "true"]:
#     logging.info("Input - SKU present (True) = file -> {0}".format(args.s.upper()))
# elif args.y in [False, "False", "false"]:
#     logging.info("Input - SKU present (False) = file -> {0}".format(args.s.upper()))
# else:
#     logging.info("Invalid input - {0} - for 'SKU present? (True/False)'".format(args.y))
logging.info("Input - SKU present = file -> {0}".format(args.s.upper()))
logging.info("Input - Name of charter = {0}".format(args.c.lower()))
logging.info("Input - Test plan name = {0}".format(args.tp.lower()))
logging.info("Input - Test plan to be manually uploaded? (True/False) = {0}".format(args.f))
if args.f in [True, "True", "true"]:
    logging.info("Input - Test plan file (True) = file -> {0}".format(args.tpf))
elif args.f in [False, "False", "false"]:
    logging.info("Input - Test plan file (False) = file -> {0}".format(args.tpf))
else:
    logging.info("Invalid input - {0} - for 'Test plan manually uploaded? (True/False)'".format(args.f))
logging.info("Input - Threshold = {0}".format(args.t))

#accessing database name from ybo_config.json file and storing it in db_name
with open("ybo_config.json") as json_data_file:
	json_data = json.load(json_data_file)
db_name = json_data["config"]["db_name"]
logging.info("Database name = {0}".format(db_name))

mandatory_tcds = []
yielding_tcds = []
non_yielding_tcds = []
blocked_tcds = []
higher_test_coverage_level_tcds = []
server_programs = ["bios.andersonlake", "bios.fishhawkfalls", "bios.eaglestream_sapphirerapids", "ifwi.eaglestream_sapphirerapids", "bios.kaseyville", "bios.birchstream"]
#legacy silicons related to a current silicon
equivalent_programs_dictionary = {
    "bios.alderlake": [],
    "bios.raptorlake": ["bios.alderlake"],
    "bios.meteorlake": ['bios.raptorlake'],
    "bios.arrowlake": ["bios.meteorlake"],
    "bios.lunarlake": [],
    "bios.kaseyville": ["bios.birchstream"]
    # "bios.rocketlake": ["bios.cometlake", "bios.tigerlake"],
    # "bios.tigerlake": ["bios.icelake", "bios.rocketlake"]
}

table_names = {
    "bios.alderlake": "ADL",
    "bios.meteorlake": "MTL",
    "bios.rocketlake": "RKL",
    "bios.tigerlake": "TGL",
    "bios.arrowlake": "ARL",
    "bios.lunarlake": "LNL",
    "bios.raptorlake": "RPL",
    "bios.andersonlake": "ANL",
    "bios.fishhawkfalls": "FHF",
    "bios.eaglestream_sapphirerapids": "EGLSR",
    "ifwi.eaglestream_sapphirerapids": "EGLSR_IFWI",
    "ifwi.alderlake": "ADL",
    "ifwi.meteorlake": "MTL",
    "ifwi.rocketlake": "RKL",
    "ifwi.tigerlake": "TGL",
    "ifwi.arrowlake": "ARL",
    "ifwi.lunarlake": "LNL",
    "ifwi.raptorlake": "RPL",
    "bios.kasyville" : "GNR_D",
    "bios.birchstream" : "GNR_AP"
}

#searching for silicons which have sku
name = table_names[args.p.lower()]
if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI", "GNR_D", "GNR_AP"]:
    table_name = name
else:
    table_name = name +'_'+ args.s.upper()

#fetching legacy silicons
equivalent_programs = []
if args.p.lower() in equivalent_programs_dictionary.keys():
    equivalent_programs = equivalent_programs_dictionary[args.p.lower()]
test_coverage_levels = {
    "l0": "l0 check-in-ci",
    "l1": "l1 dailyci-basic-sanity",
    "l2": "l2 mandatory-bat",
    "l3": "l3 extended-bat-fv",
    "l4": "l4 extended-fv",
    "l5": "l5 per-milestone",
    "l6": "l6 change-based-validation"
}
# Uncomment below line only if "Mandatory test case level" input/dropdown is enabled on website, 
# otherwise we have default "L0" and "L1" as mandatory.
# mandatory_test_coverage_levels = [] 
mandatory_test_coverage_levels = ["l0 check-in-ci", "l1 dailyci-basic-sanity"]
# One filter on final list of TCDs to be optimized based on this. 
# We should filter out TCDs based on this list. Ex. If this list has ["L2", "L3"] then 
# we should optimize only "L2" and "L3" TCDs. 
# We should have removed mandatory TCDs ("L0" and "L1") before applying this filter.
test_coverage_levels_to_be_optimized = [] 

#bat type for client side bios and orange type for server side bios
if "bat" in args.tp.lower() or "orange" in args.tp.lower():
    test_coverage_levels_to_be_optimized = ["l2 mandatory-bat", "l3 extended-bat-fv"] 
    logging.info("Test plan input - {0} found. Hence the test coverage levels to be optimized = {1}".format(args.tp.lower(), test_coverage_levels_to_be_optimized))
#fv in client side bios and blue in server side bios
elif "fv" in args.tp.lower() or "blue" in args.tp.lower():
    test_coverage_levels_to_be_optimized = ["l3 extended-bat-fv", "l4 extended-fv", "l5 per-milestone"]
    logging.info("Test plan input - {0} found. Hence the test coverage levels to be optimized = {1}".format(args.tp.lower(), test_coverage_levels_to_be_optimized))
else:
    logging.info("Invalid input - {0} - for test plan".format(args.tp))

# Uncomment below for loop only if "Mandatory test case level" input/dropdown is enabled on website.
# for k, v in test_coverage_levels.items():
#     mandatory_test_coverage_levels.append(v)
#     if k.lower() == args.l.lower():
#         break

def main():

    logging.info("Inside main().")
    logging.info("Initializing global variables...")
    """ Variable declaration """
    global args, mandatory_tcds, yielding_tcds, non_yielding_tcds, blocked_tcds, equivalent_programs, equivalent_programs_dictionary, \
           test_coverage_levels, test_coverage_levels_to_be_optimized, mandatory_test_coverage_levels, db_name, server_programs,higher_test_coverage_level_tcds
    logging.info("Global variables initialized.")
    unique_tcd_data = []
    """ Inputs """ 
    logging.info("Inputs in main()...")
    program_name = args.p.lower()
    logging.info("Input - Program name = {0}".format(program_name))
    logging.info("Input - SKU present = file -> {0}".format(args.s.upper()))
    charter = args.c.lower()
    logging.info("Input - Charter name = {0}".format(charter))
    test_plan_name = args.tp.lower()
    if test_plan_name == "bat" or test_plan_name == "orange" :
        if program_name in server_programs:
            test_plan_name = "orange"
    elif test_plan_name == "fv" or test_plan_name == "blue":
        if program_name in server_programs:
            test_plan_name = "blue"
    logging.info("Input - Test plan name = {0}".format(test_plan_name))
    manual_test_plan_file = args.f
    logging.info("Input - Test plan to be manually uploaded? (True/False) = {0}".format(manual_test_plan_file))
    if manual_test_plan_file in [True, "True", "true"]:
        test_plan_file = args.tpf
        logging.info("Input - Test plan file (True) = file -> {0}".format(test_plan_file))
    elif manual_test_plan_file in [False, "False", "false"]:
        test_plan_file = args.tpf
        logging.info("Input - Test plan file (False) = file -> {0}".format(test_plan_file))
    else:
        logging.info("Invalid input - {0} - for 'Test plan to be manually uploaded? (True/False)'".format(args.f))
    threshold = args.t
    logging.info("Input - Threshold = {0}".format(threshold))
    # tc_level = args.l.lower()
    logging.info("Completed taking inputs in main().")

    """ Getting data """

    logging.info("Getting all data based on program_name and db_name...")
    all_db_data = get_all_data(program_name, db_name, equivalent_programs, charter)
    

    logging.info("Getting all data based on program_name and db_name done.")
    print("Length of all_db_data = {0}".format(len(all_db_data)))
    #print("all_db_data")
    #print(all_db_data)


    logging.info("Getting all data as per test plan name...")
    return_list = get_all_data_as_per_test_plan_name(test_plan_name, program_name, db_name, equivalent_programs, charter)
    logging.info("Getting all data as per test plan name done.")
    unique_tcd_ids = return_list[0]
    logging.info("Received unique_tcd_ids = {0}".format(unique_tcd_ids))
        
        # For testing purpose
        # unique_tcd_ids = ['1508608189', '1508608609', '1508607484', '1508607555', '16012351903', '1508607679', '1508606695']
    print("Length of unique_tcd_ids = {0} - first call".format(len(unique_tcd_ids)))
    print("unique_tcd_ids")
    print(unique_tcd_ids)
    tp_data = return_list[1]
    logging.info("tp-data----------------------------------------------------------------{0}".format(tp_data))
                       
    print("Length of tp_data = {0}".format(len(tp_data)))
        # print("tp_data")
        # print("TP Data - 1: ******************************************************************************************")
        # print(tp_data)
    #else:
    #    logging.info("Invalid input - {0} - for 'Test plan to be uploaded manually ? (True/False)'".format(args.f))

    """ Filtering based on Mandatory Test Coverage Level """
    """ Preparing a data structure [[TCD_ID1, TCD_Title1], [TCD_ID2, TCD_Title2], ... , [TCD_ID'N', TCD_Title'N']] where N = number of unique TCD IDs """
    temp_unq_tcds = []
    
    logging.info("Filtering based on mandatory test coverage levels...")
    logging.info("Before filtering based on mandatory test coverage levels, unique_tcd_ids = {0}".format(unique_tcd_ids))
    for unq_tcd_id in unique_tcd_ids:
        # print("Current Unique TCD: {0}".format(unq_tcd_id))
        for d in tp_data:
        # for d in all_data_buckets:
            if str(d[0]) == str(unq_tcd_id):
                # print("----------------------------------------------------------------------------")
                # print("Comparison successful.")
                # print("D[0] = {0} and unq_tcd_id = {1}".format(d[0], unq_tcd_id))
                if d[2] != None and d[2].lower() in mandatory_test_coverage_levels:
                    if [str(unq_tcd_id), d[1]] not in mandatory_tcds:
                        mandatory_tcds.append([str(unq_tcd_id), d[1]])
                else:
                    if unq_tcd_id not in temp_unq_tcds:
                        temp_unq_tcds.append(unq_tcd_id)
                    if [str(unq_tcd_id), d[1]] not in unique_tcd_data:
                        unique_tcd_data.append([str(unq_tcd_id), d[1]])

            # else:
                # print("----------------------------------------------------------------------------")
                # print("Comparison unsuccessful.")
                # print("Type of D[0] is {0} and type of unq_tcd_id is {1}".format(type(d[0]), type(unq_tcd_id)))
                # print("D[0] = {0} and unq_tcd_id = {1}".format(d[0], unq_tcd_id))
                        
    unique_tcd_ids = temp_unq_tcds
    logging.info("After filtering based on mandatory test coverage levels, unique_tcd_ids = {0}".format(unique_tcd_ids))
    logging.info("Filtering based on mandatory test coverage levels done.")
    
    print("Length of unique_tcd_ids = {0} - after filtering based on mandatory test coverage levels.".format(len(unique_tcd_ids)))
    print("unique_tcd_ids")
    print(unique_tcd_ids)
    print("Length of unique_tcd_data = {0}".format(len(unique_tcd_data)))
    print("unique_tcd_data")
    print(unique_tcd_data)

    """ Filtering based on Applicable Test Coverage Levels as per Test Plan input """
    temp_unq_tcds = []
    unique_tcd_data = []
    logging.info("Filtering based on applicable test coverage levels as per test plan input...")
    logging.info("Before filtering based on applicable test coverage levels as per test plan input, unique_tcd_ids = {0}".format(unique_tcd_ids))
    for unq_tcd_id in unique_tcd_ids:
        for d in tp_data:
        # for d in all_data_buckets:
            if str(d[0]) == str(unq_tcd_id):
                if d[2]!=None and d[2].lower() not in test_coverage_levels_to_be_optimized:
                    if [str(unq_tcd_id), d[1]] not in higher_test_coverage_level_tcds:
                        higher_test_coverage_level_tcds.append([str(unq_tcd_id), d[1]])
                else:
                    if unq_tcd_id not in temp_unq_tcds:
                        temp_unq_tcds.append(unq_tcd_id)
                    if [str(unq_tcd_id), d[1]] not in unique_tcd_data:
                        unique_tcd_data.append([str(unq_tcd_id), d[1]])
                        
    unique_tcd_ids = temp_unq_tcds
    logging.info("After filtering based on applicable test coverage levels as per test plan input, unique_tcd_ids = {0}".format(unique_tcd_ids))
    logging.info("Filtering based on applicable test coverage levels as per test plan input done.")

    print("Length of unique_tcd_ids = {0} - after filtering based on applicable test coverage levels.".format(len(unique_tcd_ids)))
    print("unique_tcd_ids")
    print(unique_tcd_ids)
    print("Length of unique_tcd_data = {0}".format(len(unique_tcd_data)))
    print("unique_tcd_data")
    #print(unique_tcd_data)

    """ Preparing a data structure [[[whole data for TCD1 TR-instance1], [whole data for TCD1 TR-instance2], ... , [whole data for TCD1 TR-instance'M']], 
        [[], [], ... , []], ... , [[whole data for TCD'N' TR-instance1], [whole data for TCD'N' TR-instance2], ... , [whole data for TCD'N' TR-instance'P']]] 
        where N = number of unique TCD 
        Ultimately, we are creating N number of buckets for keeping all relevant TR instances of corresponding TCD """
    logging.info("Preparing all_data_buckets...")
    all_data_buckets = [[] for i in unique_tcd_data]

    logging.info("Preparing all_data_buckets done.")
    print("Length of all_data_buckets = {0}".format(len(all_data_buckets)))

    print(len(unique_tcd_ids))
    for i in range(len(unique_tcd_ids)):
        #print(i)
        for d in tp_data:
            if int(d[0]) == int(unique_tcd_ids[i]):
                all_data_buckets[i].append(d)
    
    """ Convert from tuple to list """
    logging.info("Preparing all_data_buckets from tuple to list...")
    for i in range(len(all_data_buckets)):
        for j in range(len(all_data_buckets[i])):
            all_data_buckets[i][j] = list(all_data_buckets[i][j])
    logging.info("Preparing all_data_buckets from tuple to list done.")
    print("------------------")
    #print(all_data_buckets)
    
    # For info purpose

    # print("All Data")
    # print(all_data_buckets[:10])
    # print("All data 0 index length = {0}".format(len(all_data_buckets[0])))

    # For info purpose
    # counter = 0
    # for i in range(len(all_data_buckets)):
    #     counter += len(all_data_buckets[i])
    # print("Counter = {0}".format(counter))
    
    
    """ Sorting Test Results """
    #logging.info("All DATA BUCKETS: {0}".format(all_data_buckets))
    logging.info("Sorting Test Results...")
    sorted_all_data_buckets = sort_test_results_by_actual_end_date(all_data_buckets,program_name)
    print("---------------------------------------------------------------------------------------")
    #print(sorted_all_data_buckets)
    logging.info("____________________________________________________________________________________________________")
    logging.info("Sorted all data buckets {0}".format(sorted_all_data_buckets))
    logging.info("Sorting Test Results Done.")

    """ Intermediate Representation """

    logging.info("Intermediate Representation...")
    # intermediate_representation = get_intermediate_representation(unique_tcd_ids, db_name, all_db_data)
    logging.info("Intermediate Representation Done.")

    """ YBO Optimization Algorithm """

    logging.info("Running YBO Optimization Algorithm...")
    return_list = ybo_optimization(program_name, sorted_all_data_buckets, unique_tcd_ids, threshold, equivalent_programs)
    logging.info("YBO Optimization Algorithm Done.")

    yielding_tcds = return_list[0]
    non_yielding_tcds = return_list[1]
    blocked_tcds = return_list[2]

    if not os.path.exists('Output_Excels'):
        os.makedirs('Output_Excels')
    
    if manual_test_plan_file in [True, "True", "true"]:
        test_plan_file_unique_tcds = []
        test_plan_file_unique_tcd_data_dictionary = {}
        if "csv" in os.path.splitext(test_plan_file)[1]:
            with open(test_plan_file) as csv_file:
                print("inside csv file")
                csv_reader = csv.reader(csv_file, delimiter=',')
                line_count = 0
                for row in csv_reader:
                    if line_count == 0:
                        print(f'Column names are {", ".join(row)}')
                        line_count += 1
                    else:
                        if str(row[0]).startswith("=HYPERLINK"):
                            print(str(row[0]).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/'))
                            print(str(row[0]).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","'))
                            if str(row[0]).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","')[0] not in test_plan_file_unique_tcds:
                                test_plan_file_unique_tcds.append(str(row[0]).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","')[0])
                                test_plan_file_unique_tcd_data_dictionary[str(row[0]).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","')[0]] = row[1]
                        else:
                            if str(row[0]) not in test_plan_file_unique_tcds:
                                test_plan_file_unique_tcds.append(str(row[0]))
                                test_plan_file_unique_tcd_data_dictionary[str(row[0])] = row[1]
                        line_count += 1
                print(f'Processed {line_count} lines.') 
        elif "xlsx" in os.path.splitext(test_plan_file)[1] or "xls" in os.path.splitext(test_plan_file)[1]:
            wb = op.load_workbook(test_plan_file)
            sheet = wb.active

            column_tcd_id = sheet["A"]
            print("column_tcd_id = {0}".format(column_tcd_id))
            if column_tcd_id[0].value not in [None, "", " "]:
                if column_tcd_id[0].value.strip() in ["TCD_ID", "tcd_id"]:
                    for i in range(1, len(column_tcd_id)):
                        print("column_tcd_id value [{0}] = {1}".format(i, column_tcd_id[i].value))
                        if str(column_tcd_id[i].value).startswith("=HYPERLINK"):
                            print(str(column_tcd_id[i].value).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/'))
                            print(str(column_tcd_id[i].value).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","'))
                            if str(column_tcd_id[i].value).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","')[0] not in test_plan_file_unique_tcds:
                                test_plan_file_unique_tcds.append(str(column_tcd_id[i].value).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","')[0])
                                test_plan_file_unique_tcd_data_dictionary[str(column_tcd_id[i].value).strip().split('=HYPERLINK("https://hsdes.intel.com/resource/')[1].split('","')[0]] = sheet.cell(row=i+1, column=2).value
                        else:
                            if str(column_tcd_id[i].value) not in test_plan_file_unique_tcds:
                                test_plan_file_unique_tcds.append(str(column_tcd_id[i].value))
                                test_plan_file_unique_tcd_data_dictionary[str(column_tcd_id[i].value)] = sheet.cell(row=i+1, column=2).value
    
        else:
            print("Invalid file extension.")
        

        temp_yielding_tcds=[]
        temp_blocked_tcds=[]
        temp_non_yielding_tcds=[]
        temp_higher_test_coverage_level_tcds=[]
        temp_mandatory_tcds=[]
        high_test_coverage_level_id=[]
        yielding_id=[]
        non_yielding_id=[]
        blocked_id=[]
        mandatory_id=[]
        temp_test_plan_unique_tcds=[]

        for i in higher_test_coverage_level_tcds:
            if i[0] not in high_test_coverage_level_id:
                high_test_coverage_level_id.append(str(i[0]))

        for i in blocked_tcds:
            if i[0] not in blocked_id:
                blocked_id.append(str(i[0]))

        for i in yielding_tcds:
            if i[0] not in yielding_id:
                yielding_id.append(str(i[0]))

        for i in non_yielding_tcds:
            if i[0] not in non_yielding_id:
                non_yielding_id.append(str(i[0]))

        for i in test_plan_file_unique_tcds:
            temp_test_plan_unique_tcds.append(i)
        test_plan_file_unique_tcds=list(set(test_plan_file_unique_tcds))
        
        print(test_plan_file_unique_tcds)
        print("-------------------------------------------------")
        print(higher_test_coverage_level_tcds)
        print("___________________________________________________")
        print(high_test_coverage_level_id)
        for i in test_plan_file_unique_tcds:
            for j in high_test_coverage_level_id:
                if i != "None" and i!= "https://hsdes.intel.com/appstore/article/" and int(i)==int(j):
                    print("higher_test_coverage_level_tcds: ",i)
                    temp_higher_test_coverage_level_tcds.append([i,test_plan_file_unique_tcd_data_dictionary[str(i)]])
                    temp_test_plan_unique_tcds.remove(i)
        
        for i in test_plan_file_unique_tcds:
            for j in blocked_id:
                if i != "None" and i!= "https://hsdes.intel.com/appstore/article/" and int(i)==int(j):
                    print("blocked_tcds: ",i)
                    temp_blocked_tcds.append([i,test_plan_file_unique_tcd_data_dictionary[str(i)]])
                    temp_test_plan_unique_tcds.remove(i)
        
        for i in test_plan_file_unique_tcds:
            for j in yielding_id:
                if i != "None" and i!= "https://hsdes.intel.com/appstore/article/" and int(i)==int(j):
                    print("yielding_tcds: ",i)
                    temp_yielding_tcds.append([i,test_plan_file_unique_tcd_data_dictionary[str(i)]])
                    temp_test_plan_unique_tcds.remove(i)

        for i in test_plan_file_unique_tcds:
            for j in non_yielding_id:
                if i != "None" and i!= "https://hsdes.intel.com/appstore/article/" and int(i)==int(j):
                    print("non_yielding_tcds: ",i)
                    temp_non_yielding_tcds.append([i,test_plan_file_unique_tcd_data_dictionary[str(i)]])
                    temp_test_plan_unique_tcds.remove(i)
        
        for i in temp_test_plan_unique_tcds:
            if i != "None" and i!= "https://hsdes.intel.com/appstore/article/":
                temp_mandatory_tcds.append([str(i),test_plan_file_unique_tcd_data_dictionary[str(i)]])

        yielding_tcds = temp_yielding_tcds
        non_yielding_tcds = temp_non_yielding_tcds
        blocked_tcds = temp_blocked_tcds
        mandatory_tcds=temp_mandatory_tcds
        higher_test_coverage_level_tcds=temp_higher_test_coverage_level_tcds


    print("Yielding TCDS: ----------------------------------------------------------------------------------------------")
    print(yielding_tcds)
    df = pd.DataFrame({'id': [l[0] for l in yielding_tcds],
                   'title': [l[1] for l in yielding_tcds]})
    df.to_excel('./Output_Excels/Yielding_TCDs.xlsx')
    print("Non-Yielding TCDS: ----------------------------------------------------------------------------------------------")
    print(non_yielding_tcds)
    df = pd.DataFrame({'id': [l[0] for l in non_yielding_tcds],
                   'title': [l[1] for l in non_yielding_tcds]})
    df.to_excel('./Output_Excels/Non_Yielding_TCDs.xlsx')
    print("Blocked TCDS: ----------------------------------------------------------------------------------------------")
    print(blocked_tcds)
    df = pd.DataFrame({'id': [l[0] for l in blocked_tcds],
                   'title': [l[1] for l in blocked_tcds]})
    df.to_excel('./Output_Excels/Blocked_TCDs.xlsx')
    print("Mandatory TCDS: ----------------------------------------------------------------------------------------------")
    print(mandatory_tcds)
    df = pd.DataFrame({'id': [l[0] for l in mandatory_tcds],
                   'title': [l[1] for l in mandatory_tcds]})
    df.to_excel('./Output_Excels/Mandatory_TCDs.xlsx')
    print("Higher Test Coverage Level TCDS: ----------------------------------------------------------------------------------------------")
    print(higher_test_coverage_level_tcds)
    df = pd.DataFrame({'id': [l[0] for l in higher_test_coverage_level_tcds],
                   'title': [l[1] for l in higher_test_coverage_level_tcds]})
    df.to_excel('./Output_Excels/Higher_Test_Coverage_Level_TCDs.xlsx')
    # source_folder = r"C:\Users\hdave\Downloads\YBO_Portal-main"
    # destination_folder = r"C:\Users\hdave\Downloads\YBO_Portal-main"
    # files_to_move = ['yielding_tcds.xlsx', 'non_yielding_tcds.xlsx', 'blocked_tcds.xlsx', 'mandatory_tcds.xlsx', 'higher_test_coverage_level_tcds.xlsx' ]

    # iterate files
    # for file in files_to_move:
    #     # construct full file path
    #     source = os.path.join(source_folder, file)
    #     destination = os.path.join(destination_folder, file)
    #     # move file
    #     shutil.move(source, destination)

    logging.info("Yielding TCDs list: {0}".format(yielding_tcds))
    logging.info("Non Yielding TCDs list: {0}".format(non_yielding_tcds))
    logging.info("Blocked TCDs list: {0}".format(blocked_tcds))
    logging.info("Mandatory TCDs list: {0}".format(mandatory_tcds))
    logging.info("Higher Test Coverage Level TCDs list: {0}".format(higher_test_coverage_level_tcds))

    return [yielding_tcds, non_yielding_tcds, blocked_tcds, mandatory_tcds, higher_test_coverage_level_tcds]

def get_all_data(program_name, db_name, equivalent_programs, charter):
    
    """ Function to get whole db data for a given program """

    logging.info("Inside get_all_data function.")
    logging.info("Arguments -> ")
    logging.info("Argument - program_name = {0}".format(program_name))
    logging.info("Argument - db_name = {0}".format(db_name))
    logging.info("Argument - equivalent_programs = {0}".format(equivalent_programs))
    logging.info("Argument - charter = {0}".format(charter))
    all_db_data = []

    db = sqlite3.connect("./DB/" + db_name)
    print(db)
    c = db.cursor()

#appending equivalent_prog and program name to new list called release_affected
#extend method iterates for the given list and appends the content to the previous list
    release_affected = [program_name]
    release_affected.extend(equivalent_programs)
    release_affected = tuple(release_affected)

    # Query to be run with or without considering equivalent programs """
    logging.info("Running query to get all data based on program_name and db_name...")
    if len(equivalent_programs) > 0:
        if charter == "bios":
            for rel_aff in release_affected:
                name = table_names[rel_aff]
                if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI"]:
                    table = name
                else:
                    table = name +'_'+ args.s.upper()
                #query = "SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Component NOT LIKE '%security%' AND Component NOT LIKE '%ifwi%'".format(table, rel_aff)
                query = "SELECT * FROM {0}".format(table)
                print("get_all_data query = {0}".format(query))
                c.execute(query)
                column_names = [description[0] for description in c.description]
                print("Column Names (Get all data) for {0}".format(rel_aff))
                print(column_names)
                all_db_data.extend(c.fetchall())

        elif charter == "security":
            for rel_aff in release_affected:
                name = table_names[rel_aff]
                if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI"]:
                    table = name
                else:
                    table = name +'_'+ args.s.upper()
                #query = "SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Component LIKE '%security%'".format(table, rel_aff)
                query = "SELECT * FROM {0}".format(table)
                print("get_all_data query = {0}".format(query))
                c.execute(query)
                column_names = [description[0] for description in c.description]
                print("Column Names (Get all data) for {0}".format(rel_aff))
                print(column_names)
                all_db_data.extend(c.fetchall())

        elif charter == "ifwi":
            for rel_aff in release_affected:
                name = table_names[rel_aff]
                if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI"]:
                    table = name
                else:
                    table = name +'_'+ args.s.upper()
                #query = "SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Component LIKE '%ifwi%'".format(table, rel_aff)
                query = "SELECT * FROM {0}".format(table)
                print("get_all_data query = {0}".format(query))
                c.execute(query)
                column_names = [description[0] for description in c.description]
                print("Column Names (Get all data) for {0}".format(rel_aff))
                print(column_names)
                all_db_data.extend(c.fetchall())
        else:
            print("Invalid value for charter - {0}".format(charter))
    else:
        if charter == "bios":
            #c.execute("SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Component NOT LIKE '%security%' AND Component NOT LIKE '%ifwi%'".format(table_name, program_name))
            c.execute("SELECT * FROM {0}".format(table_name))
        elif charter == "security":
            c.execute("SELECT * FROM {0}".format(table_name))
            #c.execute("SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Component LIKE '%security%'".format(table_name, program_name))
        elif charter == "ifwi":
            c.execute("SELECT * FROM {0}".format(table_name))
            #c.execute("SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Component LIKE '%ifwi%'".format(table_name, program_name))
        else:
            print("Invalid value for charter - {0}".format(charter))

        column_names = [description[0] for description in c.description]
        print("Column Names (Get all data) for {0}".format(program_name))
        print(column_names)
        all_db_data = c.fetchall()

    logging.info("Query execution to get all data based on program_name and db_name done.")

    logging.info("all_db_data fetch done.")
    print("Length of all_db_data in get_all_data function = {0}".format(len(all_db_data)))
    # print("all_db_data")
    # print(all_db_data)

    """Obtain all the required columns"""

    # test_case_definition_id = ybo_cf_db.get_test_case_definition_id(all_db_data, column_names)
    # test_case_definition_title = ybo_cf_db.get_test_case_definition_title(all_db_data, column_names)
    # test_coverage_level = ybo_cf_db.get_test_coverage_level(all_db_data, column_names)
    # test_case_id = ybo_cf_db.get_test_case_id(all_db_data, column_names)
    # test_case_title = ybo_cf_db.get_test_case_title(all_db_data, column_names)
    # test_result_id = ybo_cf_db.get_test_result_id(all_db_data, column_names)
    # status_reason = ybo_cf_db.get_status_reason(all_db_data, column_names)
    # actual_end = ybo_cf_db.get_actual_end(all_db_data, column_names)
    # release_affected = ybo_cf_db.get_release_affected(all_db_data, column_names)
    # test_plan = ybo_cf_db.get_test_plan(all_db_data, column_names)
    # test_cycle = ybo_cf_db.get_test_cycle(all_db_data, column_names)

    return all_db_data

def get_all_data_as_per_test_plan_name(test_plan_name, program_name, db_name, equivalent_programs, charter):
    
    """ Get all data from db as per given test plan name """

    logging.info("Inside get_all_data_as_per_test_plan_name function.")
    logging.info("Arguments -> ")
    logging.info("Argument - test_plan_name = {0}".format(test_plan_name))
    logging.info("Argument - program_name = {0}".format(program_name))
    logging.info("Argument - sku name = {0}".format(args.s.upper()))
    logging.info("Argument - db_name = {0}".format(db_name))
    logging.info("Argument - equivalent_programs = {0}".format(equivalent_programs))
    logging.info("Argument - charter = {0}".format(charter))

    db = sqlite3.connect("./DB/" + db_name)
    c = db.cursor()

    tp_data = []
    release_affected = [program_name]
    release_affected.extend(equivalent_programs)
    release_affected = tuple(release_affected)

    logging.info("Running query to get all data based on test_plan_name, program_name and db_name...")
    if len(equivalent_programs) > 0:
        if charter == "bios":
            for rel_aff in release_affected:
                name = table_names[rel_aff]
                if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI"]:
                    table = name
                else:
                    table = name +'_'+ args.s.upper()               
                #query = "SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Test_Plan LIKE '%{2}%' AND Component NOT LIKE '%security%' AND Component NOT LIKE '%ifwi%'".format(table, rel_aff, test_plan_name)
                query = "SELECT * FROM {0} WHERE Test_Plan LIKE '%{1}%'".format(table,test_plan_name)
                print("get_all_data_as_per_test_plan_name query = {0}".format(query))
                c.execute(query)
                tp_column_names = [description[0] for description in c.description]
                print("Column Names after filtering based on test plan for {0}".format(rel_aff))
                print(tp_column_names)
                tp_data.extend(c.fetchall())
        elif charter == "security":
            for rel_aff in release_affected:
                name = table_names[rel_aff]
                if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI"]:
                    table = name
                else:
                    table = name +'_'+ args.s.upper()               
                query = "SELECT * from {0} where Test_Plan LIKE '%{1}%'".format(table,test_plan_name)
                #query = "SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Test_Plan LIKE '%{2}%' AND Component LIKE '%security%'".format(table, rel_aff, test_plan_name)
                print("get_all_data_as_per_test_plan_name query = {0}".format(query))
                c.execute(query)
                tp_column_names = [description[0] for description in c.description]
                print("Column Names after filtering based on test plan for {0}".format(rel_aff))
                print(tp_column_names)
                tp_data.extend(c.fetchall())
        elif charter == "ifwi":
            for rel_aff in release_affected:
                name = table_names[rel_aff]
                if name in ["ANL", "FHF", "EGLSR", "EGLSR_IFWI"]:
                    table = name
                else:
                    table = name +'_'+ args.s.upper()              
                query = "SELECT * from {0} where Test_Plan LIKE '%{1}%'".format(table,test_plan_name)
                #query = "SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Test_Plan LIKE '%{2}%' AND Component LIKE '%ifwi%'".format(table, rel_aff, test_plan_name)
                print("get_all_data_as_per_test_plan_name query = {0}".format(query))
                c.execute(query)
                tp_column_names = [description[0] for description in c.description]
                print("Column Names after filtering based on test plan for {0}".format(rel_aff))
                print(tp_column_names)
                tp_data.extend(c.fetchall())
        else:
            print("Invalid value for charter - {0}".format(charter))
    else:
        if charter == "bios":
            c.execute("SELECT * from {0} WHERE Test_Plan LIKE '%{1}%'".format(table_name,test_plan_name))
            #c.execute("SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Test_Plan LIKE '%{2}%' AND Component NOT LIKE '%security%' AND Component NOT LIKE '%ifwi%'".format(table_name, program_name, test_plan_name))
        elif charter == "security":
            c.execute("SELECT * from {0} WHERE Test_Plan LIKE '%{1}%'".format(table_name,test_plan_name))
            #c.execute("SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Test_Plan LIKE '%{2}%' AND Component LIKE '%security%'".format(table_name, program_name, test_plan_name))
        elif charter == "ifwi":
            c.execute("SELECT * from {0} WHERE Test_Plan LIKE '%{1}%'".format(table_name,test_plan_name))
            #c.execute("SELECT * FROM {0} WHERE Release_Affected = '{1}' AND Test_Plan LIKE '%{2}%' AND Component LIKE '%ifwi%'".format(table_name, program_name, test_plan_name))
        else:
            print("Invalid value for charter - {0}".format(charter))

        tp_column_names = [description[0] for description in c.description]
        print("Column Names after filtering based on test plan for {0}".format(program_name))
        print(tp_column_names)
        tp_data = c.fetchall()
        
    
    logging.info("Query execution to get all data based on test_plan_name, program_name and db_name done.")
    
    logging.info("tp_data fetch done.")
    print("Length of tp_data in get_all_data_as_per_test_plan_name function = {0}".format(len(tp_data)))
    print("tp_data")
    # print(tp_data)

    """ Obtain all the required columns """

    logging.info("Obtaining all the required columns...")
    tp_test_case_definition_id = ybo_cf_db.get_test_case_definition_id(tp_data, tp_column_names)
    # tp_test_case_definition_title = ybo_cf_db.get_test_case_definition_title(tp_data, tp_column_names)
    # tp_test_coverage_level = ybo_cf_db.get_test_coverage_level(tp_data, tp_column_names)
    # tp_test_case_id = ybo_cf_db.get_test_case_id(tp_data, tp_column_names)
    # tp_test_case_title = ybo_cf_db.get_test_case_title(tp_data, tp_column_names)
    # tp_test_result_id = ybo_cf_db.get_test_result_id(tp_data, tp_column_names)
    # tp_status_reason = ybo_cf_db.get_status_reason(tp_data, tp_column_names)
    # tp_actual_end = ybo_cf_db.get_actual_end(tp_data, tp_column_names)
    # tp_release_affected = ybo_cf_db.get_release_affected(tp_data, tp_column_names)
    # tp_test_plan = ybo_cf_db.get_test_plan(tp_data, tp_column_names)
    # tp_test_cycle = ybo_cf_db.get_test_cycle(tp_data, tp_column_names)
    logging.info("Obtaining all the required columns done.")

    """ Getting unique TCD IDs """
    logging.info("Getting unique_tcd_ids...")
    unique_tcd_ids = list(set(tp_test_case_definition_id))
    logging.info("Getting unique_tcd_ids done.")

    print("Length of unique_tcd_ids in get_all_data_as_per_test_plan_name function = {0}".format(len(unique_tcd_ids)))
    print("unique_tcd_ids")
    print(unique_tcd_ids)
    
    return [unique_tcd_ids, tp_data]

# def test_plan_names_for_dropdown(program_name, db_name):
    
    # """ Function to get names of test plans for given program """
    
    # db = sqlite3.connect("./DB/" + db_name)
    # c = db.cursor()

    # c.execute("SELECT DISTINCT Test_Plan FROM TestCases where Release_Affected = {0}".format(program_name))

    # column_names = [description[0] for description in c.description]
    # print("Column Names (Test Plan names for dropdown)")
    # print(column_names)
    # data = c.fetchall()
    # print(data)

    # return data[0]

def sort_test_results_by_actual_end_date(all_data_buckets,program_name):

    # temp_data_unsorted = [[['1508608609', 'SMBI012 - Type 12 - System Configuration Options', 'L0 Check-in CI', '1508777233', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '1508777310', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP [TC ID 1508777233]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.', '2020ww07.2'], 
    # ['1508608609', 'SMBI012 - Type 12 - System Configuration Options', 'L0 Check-in CI', '1508777233', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '16012519601', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP [TC ID 1508777233]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.Chaitra_Blue', '2021ww05.4'], 
    # ['1508608609', 'SMBI012 - Type 12 - System Configuration Options', 'L0 Check-in CI', '1508777233', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '16012550162', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP [TC ID 1508777233]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.Blue_FSP_2107', '2021ww14.6'], 
    # ['1508608609', 'SMBI012 - Type 12 - System Configuration Options', 'L0 Check-in CI', '1508777233', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '16012709900', 'SMBI012 - Type 12 - System Configuration Options_bios.andersonlake-RVP [TC ID 1508777233]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2109_Blue', '2021ww14.5']], 
    # [['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16012519477', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.Chaitra_Blue', ''], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16012550124', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.Blue_FSP_2107', '2021ww12.5'], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16012709872', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2109_Blue', '2021ww09.3'], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16012917573', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2113_Blue_chaitr1', ''], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16012944996', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2111_Blue_chaitr1', '2021ww06.4'], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16013064154', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.FSP_ANL_2115_Blue_Chaitr1', ''], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16013363789', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2123_Blue_chaitr1', '2021ww09.5'], 
    # ['1508607484', 'Trident Acres ping loss test', 'L2 Mandatory-BAT', '16012519451', 'Trident Acres ping loss test_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue', '16013368992', 'Trident Acres ping loss test_bios.andersonlake-RVP [TC ID 16012519451]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.Blue_ANL_FSP_2119_chaitr1', '2021ww21.6']], 
    # [['1508610525', 'L3 Extended-BAT-FV', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout', '1508777347', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '1508777470', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP [TC ID 1508777347]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.', '2021ww04.5'],
    # ['1508610525', 'L3 Extended-BAT-FV', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout', '1508777347', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '16012519680', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP [TC ID 1508777347]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.Chaitra_Blue', '2021ww16.2'],
    # ['1508610525', 'L3 Extended-BAT-FV', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout', '1508777347', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '16012550283', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP [TC ID 1508777347]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.Blue_FSP_2107', '2020ww52.5'],
    # ['1508610525', 'L3 Extended-BAT-FV', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout', '1508777347', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Blue,ANL_Orange', '16012710010', '[Post-Si] To verify FSP shall configure the PCI Express Completion timeout_bios.andersonlake-RVP [TC ID 1508777347]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2109_Blue', '2021ww09.1']], 
    # [['1508607555', 'L2 Mandatory-BAT', 'Cold Reset from UEFI Shell - 3 Cycles', '1508777498', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '1508777602', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP [TC ID 1508777498]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.', '2021ww16.4'],
    # ['1508607555', 'L2 Mandatory-BAT', 'Cold Reset from UEFI Shell - 3 Cycles', '1508777498', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '16012709668', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP [TC ID 1508777498]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2109_Orange', '2021ww17.5'],
    # ['1508607555', 'L2 Mandatory-BAT', 'Cold Reset from UEFI Shell - 3 Cycles', '1508777498', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '16013205166', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP [TC ID 1508777498]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.Orange_ANL_FSP_2119_chaitr1', '2021ww08.4'],
    # ['1508607555', 'L2 Mandatory-BAT', 'Cold Reset from UEFI Shell - 3 Cycles', '1508777498', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '16013362946', 'Cold Reset from UEFI Shell - 3 Cycles_bios.andersonlake-RVP [TC ID 1508777498]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2123_Orange_chaitr1', '']], 
    # [['16012351903', 'L2 Mandatory-BAT', '[Post-Si] To verify enumerated red core count', '1508777257', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '1508777365', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP [TC ID 1508777257]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.', ''],
    # ['16012351903', 'L2 Mandatory-BAT', '[Post-Si] To verify enumerated red core count', '1508777257', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '16012519352', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP [TC ID 1508777257]', 'open.not_run', 'bios.andersonlake.PV.Server-BIOS.Chaitra_Orange', ''],
    # ['16012351903', 'L2 Mandatory-BAT', '[Post-Si] To verify enumerated red core count', '1508777257', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '16012709639', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP [TC ID 1508777257]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.ANL_FSP_2109_Orange', '2021ww17.4'],
    # ['16012351903', 'L2 Mandatory-BAT', '[Post-Si] To verify enumerated red core count', '1508777257', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP', 'bios.andersonlake', 'ANL_Orange', '16013205272', '[Post-Si] To verify enumerated red core count_bios.andersonlake-RVP [TC ID 1508777257]', 'complete.pass', 'bios.andersonlake.PV.Server-BIOS.Orange_ANL_FSP_2119_chaitr1', '2021ww14.3']]]

    # For temporary data testing 

    # for d in range(len(temp_data_unsorted)):
    #     date_raw = []
    #     date_raw_sorted = []
    #     temp_d = []
    #     for t in range(len(temp_data_unsorted[d])):
    #         if temp_data_unsorted[d][t][-1] not in ["", " ", None]:
    #             date_raw.append(temp_data_unsorted[d][t][-1])
    #     print("Dates Raw Before Sorting")
    #     print(date_raw)
    #     date_raw_sorted = sorted(date_raw)
    #     date_raw_sorted.reverse()
    #     print("Dates Raw After Sorting")
    #     print(date_raw_sorted)
    #     for l in range(len(date_raw_sorted)):
    #         for mk in range(len(temp_data_unsorted[d])):
    #             if date_raw_sorted[l] == temp_data_unsorted[d][mk][-1]:
    #                 temp_d.append(temp_data_unsorted[d][mk])
    #     for t in range(len(temp_data_unsorted[d])):
    #         if temp_data_unsorted[d][t][-1] in ["", " ", None]:
    #             temp_d.append(temp_data_unsorted[d][t])
    #     temp_data_unsorted[d] = temp_d
    #     print("Length of temp data = {0}".format(len(temp_data_unsorted[d])))
    #     print("Temp Data After Sorting")
    #     print(temp_data_unsorted[d])

    # For whole data (Actual use)

#len of all_data_buckets is number of unique tc_id
    eq_program = False
    if len(equivalent_programs) > 0:
        eq_program = True
    rel_aff = [program_name]
    rel_aff.extend(equivalent_programs)
    rel_aff = tuple(rel_aff)
    for d in range(len(all_data_buckets)):
        # print(all_data_buckets[d])
        date_raw = []
        date_raw_sorted = []
        temp_d = []
        #for a particular tcid a list is built to collect all the dates 
        for t in range(len(all_data_buckets[d])):
            # print(all_data_buckets[d][t])
            if all_data_buckets[d][t][11] not in ["", " ", None]:
                date_raw.append(all_data_buckets[d][t][11])
        print(date_raw)
        if len(date_raw):
            # print(date_raw)
            # print(all_data_buckets[d][t][0])
            logging.info("Date sort for {0} - before sorting: {1}".format(all_data_buckets[d][t][0], date_raw))
            date_raw_sorted = sorted(date_raw)
            date_raw_sorted.reverse()
            print("Dates Raw After Sorting")
            logging.info("Date sort for {0} - after sorting: {1}".format(all_data_buckets[d][t][0], date_raw_sorted))
            print(date_raw_sorted)
            date_raw_sorted=set(date_raw_sorted)
            date_raw_sorted=list(date_raw_sorted)
            date_raw_sorted=sorted(date_raw_sorted)
            date_raw_sorted.reverse()
            for l in range(len(date_raw_sorted)):
                for mk in range(len(all_data_buckets[d])):
                    #sorting done and matching and appending entire single data in temp list
                    if date_raw_sorted[l] == all_data_buckets[d][mk][11]:
                        #if all_data_buckets[d][mk] not in temp_d:
                        temp_d.append(all_data_buckets[d][mk])
            all_data_buckets[d] = temp_d
            #logging.info("------------------------ALL DATA BUCKETS for: {0}".format(all_data_buckets[d]))
            date_entry={}
            for i in range(len(all_data_buckets[d])):
                if all_data_buckets[d][i][11] not in date_entry.keys():
                    date_entry[all_data_buckets[d][i][11]]=1
                else:
                    date_entry[all_data_buckets[d][i][11]]+=1
            #logging.info("------------------------Dictionary: {0}".format(date_entry))
            counter=0
            temp_d=[]
            if eq_program:
                for l in range(len(date_raw_sorted)):
                    result="complete.pass"
                    for j in rel_aff:
                        match=False
                        for mk in range(date_entry[date_raw_sorted[l]]):
                            if all_data_buckets[d][counter+mk][5].split(".")[1]==j.split(".")[1]:
                                match=True
                                rel_list=all_data_buckets[d].index(all_data_buckets[d][counter+mk])
                                if all_data_buckets[d][counter+mk][9]=="blocked.other":
                                    result="blocked.other"
                                    break
                                elif all_data_buckets[d][counter+mk][9]=="complete.fail":
                                    result="complete.fail"
                        if match==True:
                            all_data_buckets[d][rel_list][9]=result
                            temp_d.append(all_data_buckets[d][rel_list])
                    counter=counter+date_entry[date_raw_sorted[l]]
            else:
                for l in range(len(date_raw_sorted)):
                    result="complete.pass"
                    for mk in range(date_entry[date_raw_sorted[l]]):
                        if all_data_buckets[d][counter+mk][9]=="blocked.other":
                            result="blocked.other"
                            break
                        elif all_data_buckets[d][counter+mk][9]=="complete.fail":
                            result="complete.fail"
                    counter=counter+date_entry[date_raw_sorted[l]]-1
                    all_data_buckets[d][counter][9]=result
                    temp_d.append(all_data_buckets[d][counter])
                    counter+=1
                


            #appending data which don't have dates at the end of temp list
            for t in range(len(all_data_buckets[d])):
                if all_data_buckets[d][t][11] in ["", " ", None]:
                    temp_d.append(all_data_buckets[d][t])
            all_data_buckets[d] = temp_d
            print("Length of temp data = {0}".format(len(all_data_buckets[d])))
            print("Temp Data After Sorting")
            print(all_data_buckets[d])
        else:
            break

    """ Sorting Dates in standard way """
    
    # For temporary data testing
    
    # for td in range(len(temp_data_unsorted)):
    #     temp_d = []
    #     dates = {} # {"YYYY'ww'ww.wd": "year-ww-wd" format | Ex. {"2021ww03.5": "2021-03-5"}
    #     for rw in range(len(temp_data_unsorted[td])):
    #         if temp_data_unsorted[td][rw][-1] not in ["", " ", None]:
    #             temp_list = temp_data_unsorted[td][rw][-1].split("ww")
    #             year = temp_list[0]
    #             ww = temp_list[1].split(".")[0]
    #             wd = temp_list[1].split(".")[1]
    #             print("year = {0}".format(year))
    #             print("ww = {0}".format(ww))
    #             print("wd = {0}".format(wd))
    #             date_string = year + "-" + ww + "-" + wd
    #             dates[temp_data_unsorted[td][rw][-1]] == date_string

    #     date_time_format_final = {} # dictionary {"2021ww03.5": "2021-19-01 000:000:000"}
    #     for ww_date_format, each_date in dates.items():
    #         if each_date:
    #             r = datetime.strptime(each_date, "%Y-%W-%w") 
    #             time_delta = timedelta(days=7)
    #             r = r - time_delta
    #             print(r)
    #             date_time_format_final[ww_date_format] = r
    
    #     date_time_format_final = {k: v for k, v in sorted(date_time_format_final.items(), key=lambda item: item[1])}
    #     sorted_ww_format_dates = [k for k, v in date_time_format_final.items()]
    #     sorted_ww_format_dates.reverse()
    #     print("Final Dates: {0}".format(date_time_format_final))

    #     for l in range(len(sorted_ww_format_dates)):
    #         for mp in range(len(temp_data_unsorted[td])):
    #             if sorted_ww_format_dates[l] == temp_data_unsorted[td][mp][-1]:
    #                 temp_d.append(temp_data_unsorted[td][mp])
    #     for tx in range(len(temp_data_unsorted[td])):
    #         if temp_data_unsorted[td][tx][-1] in ["", " ", None]:
    #             temp_d.append(temp_data_unsorted[td][tx])
    #     temp_data_unsorted[td] = temp_d
    #     print("Length of temp data = {0}".format(len(temp_data_unsorted[td])))
    #     print("Temp Data After Sorting")
    #     print(temp_data_unsorted[td])

    # # For whole data (actual use) 
    # for td in range(len(all_data_buckets)):
    #     temp_d = []
    #     dates = {} # {"YYYY'ww'ww.wd": "year-ww-wd" format | Ex. {"2021ww03.5": "2021-03-5"}
    #     for rw in range(len(all_data_buckets[td])):
    #         if all_data_buckets[td][rw][-1] not in ["", " ", None]:
    #             temp_list = all_data_buckets[td][rw][-1].split("ww")
    #             year = temp_list[0]
    #             ww = temp_list[1].split(".")[0]
    #             wd = temp_list[1].split(".")[1]
    #             print("year = {0}".format(year))
    #             print("ww = {0}".format(ww))
    #             print("wd = {0}".format(wd))
    #             date_string = year + "-" + ww + "-" + wd
    #             dates[all_data_buckets[td][rw][-1]] == date_string

    #     date_time_format_final = {} # dictionary {"2021ww03.5": "2021-19-01 000:000:000"}
    #     for ww_date_format, each_date in dates.items():
    #         if each_date:
    #             r = datetime.strptime(each_date, "%Y-%W-%w") 
    #             time_delta = timedelta(days=7)
    #             r = r - time_delta
    #             print(r)
    #             date_time_format_final[ww_date_format] = r
    
    #     date_time_format_final = {k: v for k, v in sorted(date_time_format_final.items(), key=lambda item: item[1])}
    #     sorted_ww_format_dates = [k for k, v in date_time_format_final.items()]
    #     sorted_ww_format_dates.reverse()
    #     print("Final Dates: {0}".format(date_time_format_final))

    #     for l in range(len(sorted_ww_format_dates)):
    #         for mp in range(len(all_data_buckets[td])):
    #             if sorted_ww_format_dates[l] == all_data_buckets[td][mp][-1]:
    #                 temp_d.append(all_data_buckets[td][mp])
    #     for tx in range(len(all_data_buckets[td])):
    #         if all_data_buckets[td][tx][-1] in ["", " ", None]:
    #             temp_d.append(all_data_buckets[td][tx])
    #     all_data_buckets[td] = temp_d
    #     print("Length of temp data = {0}".format(len(all_data_buckets[td])))
    #     print("Temp Data After Sorting")
    #     print(all_data_buckets[td])

    return all_data_buckets

def get_intermediate_representation(unique_tcd_ids, db_name, all_db_data):
    
    db = sqlite3.connect("./DB/" + db_name)
    c = db.cursor()

    intermediate_representation = {}

    """ Intermediate Representation """

    # Get unique Test Cycles 
    unique_test_cycles = []
    test_cycle_end_date = []
    
    c.execute("SELECT DISTINCT Test_Cycle FROM {0}".format(table_name))

    tc_column_names = [description[0] for description in c.description]
    print("Column Names (getting unique test cycles)")
    print(tc_column_names)
    tc_data = c.fetchall()
    print("Test Cycle Data")
    print(tc_data)
    unique_test_cycles = [i[0] for i in tc_data]
    print("Unique Test Cycle Data")
    print(unique_test_cycles)

    """ Sort according to actual_finished_ww """
    # unique_test_cycles_with_date_dict = {}
    # for tcycle_title in unique_test_cycles:
    #     eql_test_cycle = "\"eql\":\"select id, title, milestone.actual_finished_ww where " \
    #                     "tenant='central_firmware' and subject='milestone' " \
    #                     "and title = '{0}'\"".format(tcycle_title)

    #     payload_test_cycle = """{""" + """{0}""".format(eql_test_cycle) + """}"""
    #     response_eql_tcs = requests.post(base_url, verify = False, auth = HTTPBasicAuth("kdeshmuk",
    #                 "L7MCEtEu4TpoWdJ5cDfGcl45vHZrIhak2Q8xmGSuOuFezJSQ="), headers = headers,
    #                             data = payload_test_cycle)
    #     resultant = response_eql_tcs.json()['data']
    #     print("Test Cycle Resultant")
    #     print(resultant)
    #     test_cycle_actual_finished_ww = resultant[0]["milestone.actual_finished_ww"]
    #     test_cycle_end_date.append(str(test_cycle_actual_finished_ww))
    #     unique_test_cycles_with_date_dict[tcycle_title] = test_cycle_actual_finished_ww
        
    # # Sort dictionary {"test_cycle": "actual_finished_ww"}
    # unique_test_cycles_with_date_dict = {k: v for k, v in sorted(unique_test_cycles_with_date_dict.items(), key=lambda item: item[1])}
    # unique_test_cycles = [k for k, v in unique_test_cycles_with_date_dict.items()]
    # unique_test_cycles.reverse()

    """ Sort according to last executed test result """
    unique_test_cycles_with_date_dict = {}
    for tcycle_title in unique_test_cycles:
        
        c.execute("SELECT End_Date FROM {0} where Test_Cycle = {1}".format(table_name, tcycle_title))

        tcycle_column_names_tr = [description[0] for description in c.description]
        print("Column Names (getting all data for a test cycle)")
        print(tcycle_column_names_tr)
        tcycle_data_tr = c.fetchall()
        print("Test Result End Date Data for a specific Test Cycle")
        print(tcycle_data_tr)

        temp_dates = []
        test_cycles_without_end_date_data = []
        for tr_end_date in tcycle_data_tr:
            if tr_end_date[0].strip() != '' and tr_end_date[0] != None:
                temp_dates.append(str(tr_end_date[0]))
        if len(temp_dates) > 0:
            temp_dates = list(set(temp_dates))
            temp_dates = sorted(temp_dates)
            latest_date = temp_dates[-1]
            unique_test_cycles_with_date_dict[tcycle_title] = latest_date
        else:
            test_cycles_without_end_date_data.append(tcycle_title)

    # Sort dictionary {"test_cycle": "test_result_end_date"}
    unique_test_cycles_with_date_dict = {k: v for k, v in sorted(unique_test_cycles_with_date_dict.items(), key=lambda item: item[1])}
    unique_test_cycles = [k for k, v in unique_test_cycles_with_date_dict.items()]
    unique_test_cycles.reverse()
    for e in test_cycles_without_end_date_data:
        unique_test_cycles.append(e)

    # Intermediate Representation (Populating dictionary)
    for unq_tcd in unique_tcd_ids:
        intermediate_representation[unq_tcd] = [[] for i in range(len(unique_test_cycles))]
        for row in all_db_data:
            for r in row:
                if r[0] == unq_tcd:
                    tcycle = r[10]
                    if len(intermediate_representation[unq_tcd][unique_test_cycles.index(tcycle)]) == 0:
                        intermediate_representation[unq_tcd][unique_test_cycles.index(tcycle)].append(str(r[9]))

    return intermediate_representation

def is_date_within_two_qtrs(date):
    # date = "2021ww25.3"
    dates = {}
    # print(dates)
    if (bool(dates) == True):
        temp_list = date.split("ww")
        # print(temp_list)
        year = temp_list[0]
        ww = temp_list[1].split(".")[0]
        wd = temp_list[1].split(".")[1]
        # print("year = {0}".format(year))
        # print("ww = {0}".format(ww))
        # print("wd = {0}".format(wd))
        date_string = year + "-" + ww + "-" + wd
        dates[date] = date_string

        date_time_format_final = {} # dictionary {"2021ww03.5": "2021-19-01 000:000:000"}
        for ww_date_format, each_date in dates.items():
            if each_date:
                r = datetime.strptime(each_date, "%Y-%W-%w") 
                time_delta = timedelta(days=7)
                r = r - time_delta
                print(r)
                date_time_format_final[ww_date_format] = r

        date_time_format_final = {k: v for k, v in sorted(date_time_format_final.items(), key=lambda item: item[1])}
        sorted_ww_format_dates = [k for k, v in date_time_format_final.items()]
        sorted_ww_format_dates.reverse()
        print("Final Dates: {0}".format(date_time_format_final))

        delta = datetime.now() - date_time_format_final[date]
        print(delta.days)
        if delta.days <= int(365/2):
            return True
        else:
            False
    else:
        False
        
def ybo_optimization(program_name, sorted_all_data_buckets, unique_tcd_ids, threshold, equivalent_programs):
    global yielding_tcds, non_yielding_tcds, blocked_tcds, mandatory_tcds
    eq_program = False
    if len(equivalent_programs) > 0:
        eq_program = True
    print("eq_program = {0}".format(eq_program))
    blocked_tcds = []
    yielding_tcds = []
    non_yielding_tcds = []    
    c1s = ["bios.alderlake", "bios.meteorlake", "bios.arrowlake", "bios.lunarlake", "bios.raptorlake",
           "bios.tigerlake", "bios.jasperlake", "bios.rocketlake"]

    """ Separate out blocked test cases """
    # find TCDs where last executed TR was "blocked". Add those TCDs to blocked_tcds list
    logging.info("Before preparing blocked_tcds list, unique_tcd_ids = {0}".format(unique_tcd_ids))
    temp_unq_tcds = []
    unique_tcd_data = []
    for unq_tcd_id in unique_tcd_ids:
        for d in sorted_all_data_buckets:
            # print(len(sorted_all_data_buckets))
            # print(sorted_all_data_buckets)
            if len(d):
                if str(d[0][0]) == str(unq_tcd_id):
                    if "blocked" in d[0][9].lower() or "not_run" in d[0][9].lower():
                        if [str(unq_tcd_id), d[0][1]] not in blocked_tcds:
                            blocked_tcds.append([str(unq_tcd_id), d[0][1]])
                    else:
                        temp_unq_tcds.append(unq_tcd_id)
                        if [str(unq_tcd_id), d[0][1]] not in unique_tcd_data:
                            unique_tcd_data.append([str(unq_tcd_id), d[0][1]])
            
                                      
    unique_tcd_ids = temp_unq_tcds
    logging.info("After preparing blocked_tcds list, unique_tcd_ids = {0}".format(unique_tcd_ids))
    logging.info("After preparing blocked_tcds list, blocked_tcds = {0}".format(blocked_tcds))
    # count number of test results
    for unq_tcd_id in unique_tcd_ids:
        """ Threshold criteria """
        logging.info("Currently running ybo_optimization algorithm on {0}".format(unq_tcd_id))
        current_program_tr_counter = 0
        eq_program_tr_counter = [0 for i in range(len(equivalent_programs))]
        for row in sorted_all_data_buckets:
            if len(row):
                if str(row[0][0]) == str(unq_tcd_id):
                    # Counting number of TRs
                    logging.info("Row = {0}".format(row))
                    if eq_program:
                        for row_record in row:
                            if "bios." + row_record[5].split(".")[1] in equivalent_programs:
                                eq_program_tr_counter[equivalent_programs.index("bios." + row_record[5].split(".")[1])] += 1
                            elif "bios." + row_record[5].split(".")[1] == program_name:
                            #elif program_name.split(".")[1] in row_record[5]:
                                current_program_tr_counter += 1
                    else:
                        current_program_tr_counter = len(row)
                    # Counting number of TRs done
                    logging.info("Current program test results count = {0}".format(current_program_tr_counter))
                    if current_program_tr_counter >= max(5,threshold):
                        print("current program tr counter >= threshold")
                        # Next steps
                        failed_tcd_data_current = []
                        failed_tcd_data_legacy = []
                        failed_tcd_data_current_effectiveness_logic = []
                        failure_counter_current_program = 0
                        failure_counter_legacy_program = [0 for i in range(len(equivalent_programs))]
                        failure_counter_current_program_effectiveness_logic = 0
                        failure_date_current = ""
                        failure_date_legacy = ""
                        """ Logic to find any failure in current program in last n(6) runs and increase counter by one NOTE: When we get first failure, add the data to 
                            failed_tcd_data_current and we break the loop """
                        # 1. Check if any failure in current program for last (threshold/2)(12/2 = 6) runs, keep a list failed_tcd_data_current = []
                        counter1 = 1
                        for run in range(len(row)):
                            if counter1 <= max(5,threshold):
                                if "bios."+row[run][5].split(".")[1] == program_name:
                                #if program_name.split(".")[1] in row[run][5]:
                                    counter1 += 1
                                    if "fail" in row[run][9] or "blocked" in row[run][9]:
                                        if [str(unq_tcd_id), row[0][1]] not in failed_tcd_data_current:
                                            failed_tcd_data_current.append([str(unq_tcd_id), row[0][1]])
                                            failure_date_current = str(row[run][11])
                                            failure_counter_current_program += 1
                                            break
                            else:
                                break
                        print("found {0} failures in current program".format(failure_counter_current_program))
                        logging.info("Current program failures = {0}".format(failure_counter_current_program))
                        if failure_counter_current_program > 0:
                            logging.info("Putting {0} into yielding as it has failed at least once (failure counter = {1}) in last (threshold/2) runs.".format(unq_tcd_id, failure_counter_current_program))
                            yielding_tcds.append([unq_tcd_id, row[0][1]])
                        else:
                            logging.info("{0} has passed in last (threshold/2) runs, hence going for next steps.".format(unq_tcd_id))
                            if eq_program:
                                if all(cr >= threshold for cr in eq_program_tr_counter):
                                    logging.info("Equivalent programs to be considered, going for next steps.")
                                    """ Logic to find any failure in legacy program in last n(6) runs and increase counter by one NOTE: When we get first failure, add the data to failed_tcd_data_legacy and we break the loop """
                                    counter2 = 0
                                    for run in range(len(row)):
                                        if counter2 <= max(5,threshold):
                                            if "bios." + row[run][5].split(".")[1] in equivalent_programs:
                                                counter2 += 1
                                                if "fail" in row[run][9]:
                                                    if [str(unq_tcd_id), row[0][1]] not in failed_tcd_data_legacy:
                                                        failed_tcd_data_legacy.append([str(unq_tcd_id), row[0][1]])
                                                        failure_date_legacy = row[run][11]
                                                        print(failure_date_legacy)
                                                        failure_counter_legacy_program[equivalent_programs.index("bios." + row[run][5].split(".")[1])] += 1
                                                        break
                                        else:
                                            break
                                    logging.info("Failures in legacy/equivalent programs = {0}".format(failure_counter_legacy_program))
                                    print("Found {0} failures in legacy program".format(failure_counter_legacy_program))
                                # Check if equivalent programs have any failure
                                    if any(fd > 0 for fd in failure_counter_legacy_program):
                                        logging.info("Found at least one failure in one of the legacy/equivalent programs, going for effectiveness logic.")
                                    # Effectiveness logic
                                    # if (program_name in c1s) and all(eqprog in C1S for eqprog in equivalent_programs):
                                        print("Effectiveness logic")
                                        print("program_name = {0}".format(program_name))
                                        print("row[run][5] = {0}".format(row[run][5]))
                                        if (program_name in c1s) and ("bios." + row[run][5].split(".")[1] in c1s):
                                    #if (program_name.split(".")[1] in c1s) and (row[run][5] in c1s):
                                            logging.info("Current and legacy/equivalent programs both of them are in C1S list, going for checking whether failure date in equivalent program is within 2 QTRs.")
                                            """ Logic to check if the failure_date of failed test result in legacy is within 2 QTRs """
                                            # Check if failure date <= 2 QTR
                                            if is_date_within_two_qtrs(failure_date_legacy):
                                                logging.info("Failure date in legacy/equivalent program is within 2 QTRs. Hence putting {0} in yielding list.".format(unq_tcd_id))
                                                yielding_tcds.append([unq_tcd_id, row[0][1]])
                                            else:
                                                logging.info("Failure date in legacy/equivalent program is not within 2 QTRs. Hence putting {0} in non_yielding list.".format(unq_tcd_id))
                                                non_yielding_tcds.append([unq_tcd_id, row[0][1]])
                                        else:
                                            """ Logic to find any failure in current program in last 2n(12) runs and increase counter by one NOTE: When we get first failure, add the data to 
                                                and we break the loop """
                                            # Check if any failure in current program for last (threshold)(12) runs
                                            logging.info("Both the programs (current and legacy/equivalent) are not in C1S list, going for checking any failure in last (threshold) runs.")
                                            counter3 = 1
                                            if len(row)>=2*threshold:
                                                for run in range(len(row)):
                                                    if counter3 <= 2*threshold:
                                                        if "bios." + row[run][5].split(".")[1] == program_name:
                                                    #if program_name.split(".")[1] in row[run][5]:
                                                            counter3 += 1
                                                            if "fail" in row[run][9]:
                                                                if [str(unq_tcd_id), row[0][1]] not in failed_tcd_data_current_effectiveness_logic:
                                                                    failed_tcd_data_current_effectiveness_logic.append([str(unq_tcd_id), row[0][1]])
                                                                    failure_counter_current_program_effectiveness_logic += 1
                                                                    break
                                                    else:
                                                        break
                                                print("found {0} failures in failure_counter_current_program_effectiveness_logic".format(failure_counter_current_program_effectiveness_logic))
                                                logging.info("Count of failures in last (threshold) runs (effectiveness logic) = {0}".format(failure_counter_current_program_effectiveness_logic))
                                                if failure_counter_current_program_effectiveness_logic > 0:
                                                    logging.info("Found at least one failure in last (threshold) runs, hence putting {0} in yielding list.".format(unq_tcd_id))
                                                    yielding_tcds.append([unq_tcd_id, row[0][1]])
                                                else:
                                                    logging.info("Did not found any failure in last (threshold) runs, hence putting {0} in non_yielding list.".format(unq_tcd_id))
                                                    non_yielding_tcds.append([unq_tcd_id, row[0][1]])
                                            else:
                                                 yielding_tcds.append([unq_tcd_id, row[0][1]])       
                                    else:
                                        logging.info("No failure found in legacy/equivalent programs, hence putting {0} in non_yielding list.".format(unq_tcd_id))
                                        non_yielding_tcds.append([unq_tcd_id, row[0][1]])
                                else:
                                    yielding_tcds.append([unq_tcd_id, row[0][1]])
                            else:
                                logging.info("Equivalent programs not to be considered, hence putting {0} in non_yielding list.".format(unq_tcd_id))
                                non_yielding_tcds.append([unq_tcd_id, row[0][1]])
                    else:
                        logging.info("Current programs does not have sufficient data, hence putting {0} in yielding list.".format(unq_tcd_id))
                        yielding_tcds.append([unq_tcd_id, row[0][1]])


    return [yielding_tcds, non_yielding_tcds, blocked_tcds] 

if __name__=='__main__':
    print("Running YBO.")
    logging.info("Calling main()...")
    # final_output_lists = [yielding_tcds, non_yielding_tcds, blocked_tcds, mandatory_tcds, higher_test_coverage_level_tcds]
    # final_output_lists is a list of lists
    # Each list in final_output_lists contains a list of format [[id, title], [id, title], ... , [id, title]]
    final_output_lists = main()